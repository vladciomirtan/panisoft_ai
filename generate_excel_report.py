import os
import sys
import pandas as pd
import gc
from tqdm import tqdm
from document_processor import extract_text, load_cvs, load_job_descriptions
from matcher import CVJobMatcher
from config import GEMINI_API_KEY, CV_DIR, JOB_DESCRIPTIONS_DIR, OUTPUT_DIR
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from matching_algorithm import match_cv_to_job_description

def generate_excel_report(cv_sample_size=None, job_sample_size=None):
    """
    Generate an Excel report of CV-Job matches.
    
    Args:
        cv_sample_size (int, optional): Number of CVs to use (for testing with smaller dataset)
        job_sample_size (int, optional): Number of jobs to use (for testing with smaller dataset)
    """
    try:
        # Check if API key is provided
        if not GEMINI_API_KEY:
            print("Error: GEMINI_API_KEY is not set.")
            print("Please set your API key in the config.py file or as an environment variable.")
            sys.exit(1)
        
        print("Using Gemini API")
        
        print("=" * 80)
        print("Generating Excel Report for CV-Job Matching")
        print("=" * 80)
        
        # Get the project root directory
        project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        
        # Get paths to sample files
        cv_dir = os.path.join(project_root, "panisoft_ai", "DataSet", "cv")
        job_dir = os.path.join(project_root, "panisoft_ai", "DataSet", "job_descriptions")
        
        # Load the CVs and job descriptions
        print("Loading CVs and job descriptions...")
        cv_files = [f for f in os.listdir(cv_dir) if f.endswith(('.docx', '.pdf'))]
        job_files = [f for f in os.listdir(job_dir) if f.endswith(('.docx', '.pdf'))]
        
        # If sample sizes are specified, limit the number of files
        if cv_sample_size:
            print(f"Using sample size: {cv_sample_size} CVs")
            cv_files = cv_files[:cv_sample_size]
        if job_sample_size:
            print(f"Using sample size: {job_sample_size} job descriptions")
            job_files = job_files[:job_sample_size]
        
        # Initialize matcher
        matcher = CVJobMatcher()
        
        # Create a DataFrame to store the results
        # Headers: CV, Job, Industry Score, Technical Score, Match Score, Total Score, Reasoning
        results = []
        
        # Create a lookup to convert file names to more readable names
        cv_names = {}
        for f in cv_files:
            if f.endswith('.docx'):
                cv_names[f] = f.replace('cv_', '').replace('.docx', '')
            else:
                # For PDF files, just use the filename without extension
                cv_names[f] = f.replace('.pdf', '')
        
        job_names = {}
        for f in job_files:
            if f.endswith('.docx'):
                job_names[f] = f.replace('job_description_', '').replace('.docx', '')
            else:
                # For PDF files, just use the filename without extension
                job_names[f] = f.replace('.pdf', '')
        
        # Calculate total number of matches
        total_matches = len(cv_files) * len(job_files)
        print(f"Running {total_matches} matches ({len(cv_files)} CVs x {len(job_files)} jobs)...")
        
        # Create progress bar
        progress = tqdm(total=total_matches, desc="Matching CVs with Jobs")
        
        # For each CV and job description, calculate the match
        for job_file in job_files:
            job_path = os.path.join(job_dir, job_file)
            job_content = extract_text(job_path)
            
            for cv_file in cv_files:
                cv_path = os.path.join(cv_dir, cv_file)
                cv_content = extract_text(cv_path)
                
                # Match CV with job description
                result = matcher.match(cv_content, job_content)
                
                # Store the result
                results.append({
                    'CV': cv_names[cv_file],
                    'CV_File': cv_file,
                    'Job': job_names[job_file],
                    'Job_File': job_file,
                    'Industry_Score': result.industry_knowledge_score,
                    'Technical_Score': result.technical_skills_score,
                    'Match_Score': result.job_description_match_score,
                    'Total_Score': result.total_score,
                    'Reasoning': result.reasoning[:500]  # Limit reasoning length to avoid Excel issues
                })
                
                # Update progress bar
                progress.update(1)
        
        # Close progress bar
        progress.close()
        
        # Convert to DataFrame
        df = pd.DataFrame(results)
        
        # Add percentage columns for easier reading
        df['Industry_Score_Pct'] = df['Industry_Score'] * 100
        df['Technical_Score_Pct'] = df['Technical_Score'] * 100
        df['Match_Score_Pct'] = df['Match_Score'] * 100
        df['Total_Score_Pct'] = df['Total_Score'] * 100
        
        # Create output directory if it doesn't exist
        output_dir = os.path.join(project_root, OUTPUT_DIR)
        os.makedirs(output_dir, exist_ok=True)
        
        # Create Excel file
        excel_file = os.path.join(output_dir, "cv_job_matching_results.xlsx")
        print(f"\nSaving results to {excel_file}...")
        
        # Create a Pandas Excel writer
        writer = pd.ExcelWriter(excel_file, engine='openpyxl')
        
        try:
            # Write the main results sheet
            df.to_excel(writer, sheet_name='All_Matches', index=False)
            
            # Create a sheet with the best CV for each job
            best_matches = df.loc[df.groupby('Job')['Total_Score'].idxmax()]
            best_matches = best_matches.sort_values('Total_Score', ascending=False)
            best_matches.to_excel(writer, sheet_name='Best_Match_Per_Job', index=False)
            
            # Create a sheet with the top 5 jobs for each CV
            top_jobs_per_cv = pd.DataFrame()
            for cv in df['CV'].unique():
                cv_data = df[df['CV'] == cv].sort_values('Total_Score', ascending=False).head(5)
                top_jobs_per_cv = pd.concat([top_jobs_per_cv, cv_data])
            top_jobs_per_cv.to_excel(writer, sheet_name='Top5_Jobs_Per_CV', index=False)
            
            # Create a sheet with the top 5 CVs for each job
            top_cvs_per_job = pd.DataFrame()
            for job in df['Job'].unique():
                job_data = df[df['Job'] == job].sort_values('Total_Score', ascending=False).head(5)
                top_cvs_per_job = pd.concat([top_cvs_per_job, job_data])
            top_cvs_per_job.to_excel(writer, sheet_name='Top5_CVs_Per_Job', index=False)
            
            # Save the workbook
            writer._save()
        finally:
            # Ensure writer is closed
            writer.close()
        
        # Clean up large variables to free memory
        del df, results, best_matches, top_jobs_per_cv, top_cvs_per_job
        
        print(f"Excel file created: {excel_file}")
        print("\nThe Excel file contains the following sheets:")
        print("1. All_Matches: All CV-job combinations")
        print("2. Best_Match_Per_Job: The best CV for each job")
        print("3. Top5_Jobs_Per_CV: The top 5 jobs for each CV")
        print("4. Top5_CVs_Per_Job: The top 5 CVs for each job")
        print("\nScores are provided in both decimal (0-1) and percentage (0-100%) formats.")
        print("Process completed successfully!")
    except Exception as e:
        print(f"Error generating Excel report: {str(e)}")
        raise
    finally:
        # Force garbage collection to clean up resources
        gc.collect()

def generate_excel_report_from_processed_data():
    """Generate an Excel report from processed matching data."""
    try:
        # Get list of CVs and job descriptions from directories
        cv_files = [f for f in os.listdir(CV_DIR) if f.endswith(('.docx', '.pdf'))]
        job_files = [f for f in os.listdir(JOB_DESCRIPTIONS_DIR) if f.endswith(('.docx', '.pdf'))]
        
        # Create arrays to hold data
        data = []
        
        # Process each CV and job description pair
        for cv_file in cv_files:
            # Create readable name for CV
            if cv_file.endswith('.docx'):
                cv_name = '_'.join(cv_file.split('_')[2:]).replace('.docx', '')
                cv_id = cv_file.split('_')[1]
            elif cv_file.endswith('.pdf'):
                # For PDF files, handle name extraction differently
                if cv_file.startswith('cv_'):
                    parts = cv_file.split('_')
                    if len(parts) >= 3:
                        cv_id = parts[1]
                        cv_name = '_'.join(parts[2:]).replace('.pdf', '')
                    else:
                        cv_id = parts[1].replace('.pdf', '')
                        cv_name = "Unnamed"
                else:
                    cv_id = "N/A"
                    cv_name = cv_file.replace('.pdf', '')
            
            for job_file in job_files:
                # Create readable name for job description
                if job_file.endswith('.docx'):
                    job_title = '_'.join(job_file.split('_')[3:]).replace('.docx', '')
                    job_id = job_file.split('_')[2]
                elif job_file.endswith('.pdf'):
                    # For PDF files, handle job title extraction differently
                    if job_file.startswith('job_description_'):
                        parts = job_file.split('_')
                        if len(parts) >= 4:
                            job_id = parts[2]
                            job_title = '_'.join(parts[3:]).replace('.pdf', '')
                        else:
                            job_id = parts[2].replace('.pdf', '')
                            job_title = "Untitled"
                    else:
                        job_id = "N/A"
                        job_title = job_file.replace('.pdf', '')
                
                # Get match results
                cv_path = os.path.join(CV_DIR, cv_file)
                job_path = os.path.join(JOB_DESCRIPTIONS_DIR, job_file)
                score, skills_match = match_cv_to_job_description(cv_path, job_path)
                
                # Add to data array
                data.append({
                    'CV ID': cv_id,
                    'Name': cv_name,
                    'Job ID': job_id,
                    'Job Title': job_title,
                    'Match Score': score,
                    'Skills Match': skills_match
                })
        
        # Create DataFrame from data
        df = pd.DataFrame(data)
        
        # Sort by match score (descending)
        df = df.sort_values(by='Match Score', ascending=False)
        
        # Create Excel workbook and add data
        wb = Workbook()
        ws = wb.active
        ws.title = "Matching Results"
        
        # Add header
        headers = ['CV ID', 'Name', 'Job ID', 'Job Title', 'Match Score', 'Skills Match']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # Add data
        for row_num, record in enumerate(df.itertuples(), 2):
            ws.cell(row=row_num, column=1).value = record.CV_ID
            ws.cell(row=row_num, column=2).value = record.Name
            ws.cell(row=row_num, column=3).value = record.Job_ID
            ws.cell(row=row_num, column=4).value = record.Job_Title
            ws.cell(row=row_num, column=5).value = f"{record.Match_Score:.2f}%"
            ws.cell(row=row_num, column=6).value = record.Skills_Match
        
        # Auto-adjust column width
        for column in ws.columns:
            max_length = 0
            column_name = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = max_length + 2
            ws.column_dimensions[column_name].width = adjusted_width
        
        # Create output directory if it doesn't exist
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        
        # Save workbook
        output_path = os.path.join(OUTPUT_DIR, "matching_results.xlsx")
        wb.save(output_path)
        wb.close()  # Explicitly close the workbook
        
        # Clean up large variables to free memory
        del df, data, wb, ws
        
        print(f"Excel report generated: {output_path}")
    finally:
        # Force garbage collection to clean up resources
        gc.collect()

if __name__ == "__main__":
    try:
        # Check if sample sizes are provided as command line arguments
        cv_sample_size = None
        job_sample_size = None
        
        if len(sys.argv) > 1:
            if sys.argv[1].isdigit():
                cv_sample_size = int(sys.argv[1])
            if len(sys.argv) > 2 and sys.argv[2].isdigit():
                job_sample_size = int(sys.argv[2])
        
        generate_excel_report(cv_sample_size, job_sample_size)
        generate_excel_report_from_processed_data()
    finally:
        # One final garbage collection to ensure all resources are freed
        gc.collect()
        # Exit explicitly
        sys.exit(0) 